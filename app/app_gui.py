"""ORMANET Upselling GUI application."""

from __future__ import annotations

import json
import logging
from logging.handlers import RotatingFileHandler
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import uuid

from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
IMPORT_DIR = BASE_DIR / "import"
ORDERS_DIR = IMPORT_DIR / "ORDINI"
OUTPUT_DIR = BASE_DIR / "output"
CONFIG_DIR = BASE_DIR / "config"
LOGS_DIR = BASE_DIR / "logs"

VAT_RATE = 0.22
ABSOLUTE_MIN_MARKUP = 11.0

LISTINO_MAP = {
    "LISTINO RI+10%": "RIV+10",
    "LISTINO RI": "RIV",
    "LISTINO DI": "DIST",
}

CAUSALI = ["DISPONIBILE", "IN ARRIVO", "PROGRAMMATO"]
AGGRESSIVITA_STEPS = {
    0: 0.0,
    1: 2.0,
    2: 4.0,
    3: 6.0,
}


@dataclass
class ClientInfo:
    client_id: str
    ragione_sociale: str
    listino: str
    categoria: str


@dataclass
class StockItem:
    categoria: str
    marca: str
    codice: str
    descrizione: str
    disp: float
    disp_in_arrivo: float
    giacenza: float
    data_arrivo: str
    listino_ri10: float
    listino_ri: float
    listino_di: float


@dataclass
class OrderItem:
    marca: str
    categoria: str
    codice: str
    descrizione: str
    qty: float
    prezzo_unit: float


@dataclass
class UpsellRow:
    codice: str
    descrizione: str
    qty: float
    prezzo_unit: float
    markup: float
    totale: float
    disp: float
    disponibile_dal: str | None = None


class SessionLogger:
    """Structured logging with session support."""

    def __init__(self) -> None:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        self.session_id = uuid.uuid4().hex
        self.logger = logging.getLogger("ormanets")
        self.logger.setLevel(logging.INFO)
        handler = RotatingFileHandler(LOGS_DIR / "app.log", maxBytes=500_000, backupCount=3)
        formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        handler.setFormatter(formatter)
        if not self.logger.handlers:
            self.logger.addHandler(handler)
        self.errors_path = LOGS_DIR / "errors.jsonl"

    def info(self, message: str) -> None:
        self.logger.info("[%s] %s", self.session_id, message)

    def error(self, message: str, **extra: object) -> None:
        self.logger.error("[%s] %s", self.session_id, message)
        payload = {
            "timestamp": datetime.utcnow().isoformat(),
            "session_id": self.session_id,
            "message": message,
            "extra": extra,
        }
        with self.errors_path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def normalize_text(value: str) -> str:
    if value is None:
        return ""
    text = value.upper().strip()
    text = re.sub(r"[\s/_-]+", " ", text)
    text = (
        text.replace("À", "A")
        .replace("È", "E")
        .replace("É", "E")
        .replace("Ì", "I")
        .replace("Ò", "O")
        .replace("Ù", "U")
    )
    return text


def load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_clients(path: Path, logger: SessionLogger) -> list[ClientInfo]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {name: idx for idx, name in enumerate(headers)}

    clients: list[ClientInfo] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        client_id = str(row[header_map.get("ID", 0)] or "").strip()
        ragione_sociale = str(row[header_map.get("Ragione sociale", 1)] or "").strip()
        listino = str(row[header_map.get("Listino", 0)] or "").strip()
        categoria = str(row[header_map.get("Categoria", 0)] or "").strip()
        if not client_id or not ragione_sociale:
            continue
        clients.append(ClientInfo(client_id, ragione_sociale, listino, categoria))
    logger.info("Loaded %s clients", len(clients))
    return clients


def load_stock(path: Path, logger: SessionLogger) -> dict[str, StockItem]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {name: idx for idx, name in enumerate(headers)}
    stock: dict[str, StockItem] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        codice = str(row[header_map.get("Codice", 0)] or "").strip()
        if not codice:
            continue
        stock[codice] = StockItem(
            categoria=str(row[header_map.get("Categoria", 0)] or "").strip(),
            marca=str(row[header_map.get("Marca", 0)] or "").strip(),
            codice=codice,
            descrizione=str(row[header_map.get("Descrizione", 0)] or "").strip(),
            disp=float(row[header_map.get("Disp.", 0)] or 0),
            disp_in_arrivo=float(row[header_map.get("Disp. in Arrivo", 0)] or 0),
            giacenza=float(row[header_map.get("Giacenza", 0)] or 0),
            data_arrivo=str(row[header_map.get("Data evasione/arrivo", 0)] or "").strip(),
            listino_ri10=float(row[header_map.get("Listino RI+10%", 0)] or 0),
            listino_ri=float(row[header_map.get("Listino RI", 0)] or 0),
            listino_di=float(row[header_map.get("Listino DI", 0)] or 0),
        )
    logger.info("Loaded %s stock items", len(stock))
    return stock


def load_orders(paths: list[Path], logger: SessionLogger) -> list[OrderItem]:
    items: list[OrderItem] = []
    for path in paths:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {name: idx for idx, name in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            codice = str(row[header_map.get("Cod.", 0)] or "").strip()
            if not codice:
                continue
            items.append(
                OrderItem(
                    marca=str(row[header_map.get("Marca", 0)] or "").strip(),
                    categoria=str(row[header_map.get("Categoria", 0)] or "").strip(),
                    codice=codice,
                    descrizione=str(row[header_map.get("Descrizione", 0)] or "").strip(),
                    qty=float(row[header_map.get("Qty", 0)] or 0),
                    prezzo_unit=float(row[header_map.get("Prezzo (unit, ex VAT)", 0)] or 0),
                )
            )
    logger.info("Loaded %s order items", len(items))
    return items


def map_macro_category(raw_category: str, mapping: dict, logger: SessionLogger) -> str:
    normalized = normalize_text(raw_category)
    for macro, rules in mapping.items():
        for rule in rules:
            if normalize_text(rule) in normalized:
                return macro
    token_map = {
        "BATTER": "BATTERIE",
        "CANCELL": "CANCELLERIA",
        "CARTA": "CARTA",
        "ROTOL": "ROTOLI TERMICI",
        "REMAN": "REMAN",
        "ORIG": "ORIGINALI",
        "STORAGE": "STORAGE",
        "TIMBR": "TIMBRI",
    }
    for token, macro in token_map.items():
        if token in normalized:
            return macro
    logger.error("Categoria non riconosciuta", categoria=raw_category)
    return "UNKNOWN"


def required_markup(macro: str, listino: str, sconti: dict) -> float:
    listino_key = LISTINO_MAP.get(listino.upper().strip(), "RIV")
    ric = sconti.get(macro, {}).get(listino_key, {}).get("ric", ABSOLUTE_MIN_MARKUP)
    return max(ABSOLUTE_MIN_MARKUP, float(ric))


def pick_price(item: OrderItem, stock: StockItem | None, listino: str) -> float:
    if item.prezzo_unit:
        return item.prezzo_unit
    if stock is None:
        return 0.0
    key = LISTINO_MAP.get(listino.upper().strip(), "RIV")
    if key == "RIV+10":
        return stock.listino_ri10
    if key == "DIST":
        return stock.listino_di
    return stock.listino_ri


def is_available(stock_item: StockItem, causale: str) -> tuple[bool, str | None]:
    if causale == "PROGRAMMATO":
        if stock_item.disp > 0:
            return True, None
        if stock_item.disp_in_arrivo > 0 and stock_item.data_arrivo:
            return True, stock_item.data_arrivo
        return False, None
    if causale == "IN ARRIVO":
        return stock_item.disp > 0, None
    return stock_item.disp > 0, None


def compute_upsell(
    current_items: list[OrderItem],
    historical_items: list[OrderItem],
    stock: dict[str, StockItem],
    client: ClientInfo,
    sconti: dict,
    category_map: dict,
    aggressivita: int,
    causale: str,
    logger: SessionLogger,
) -> list[UpsellRow]:
    suggestions: list[UpsellRow] = []
    historical_by_code = {item.codice: item for item in historical_items}
    current_by_code = {item.codice: item for item in current_items}

    def add_suggestion(item: OrderItem) -> None:
        if item.codice in {row.codice for row in suggestions}:
            return
        stock_item = stock.get(item.codice)
        if not stock_item:
            return
        available, available_date = is_available(stock_item, causale)
        if not available:
            return
        macro = map_macro_category(item.categoria, category_map, logger)
        if macro == "UNKNOWN":
            raise ValueError(f"Categoria non riconosciuta: {item.categoria}")
        required = required_markup(macro, client.listino, sconti)
        base_price = pick_price(item, stock_item, client.listino)
        if base_price <= 0:
            return
        cost_est = base_price / (1 + required / 100)
        max_discount = AGGRESSIVITA_STEPS.get(aggressivita, 0.0)
        discounted_price = base_price * (1 - max_discount / 100)
        min_price = cost_est * (1 + required / 100)
        final_price = max(discounted_price, min_price)
        markup = ((final_price - cost_est) / cost_est) * 100 if cost_est else required
        qty = max(1.0, item.qty)
        totale = final_price * qty
        suggestions.append(
            UpsellRow(
                codice=item.codice,
                descrizione=item.descrizione,
                qty=qty,
                prezzo_unit=round(final_price, 2),
                markup=round(markup, 2),
                totale=round(totale, 2),
                disp=stock_item.disp,
                disponibile_dal=available_date,
            )
        )

    color_tokens = ["CYAN", "MAGENTA", "YELLOW"]
    for item in current_items:
        description = normalize_text(item.descrizione)
        if any(token in description for token in color_tokens):
            for hist_item in historical_items:
                if hist_item.marca == item.marca and "BLACK" in normalize_text(hist_item.descrizione):
                    add_suggestion(hist_item)
                    break

    for item in current_items:
        if len(suggestions) >= 3:
            break
        stock_item = stock.get(item.codice)
        if not stock_item or stock_item.disp <= item.qty:
            continue
        add_suggestion(item)

    if len(suggestions) < 3:
        for hist_item in historical_items:
            if len(suggestions) >= 3:
                break
            if hist_item.codice in current_by_code:
                continue
            add_suggestion(hist_item)

    logger.info("Computed %s upsell rows", len(suggestions))
    return suggestions[:3]


def export_excel(rows: list[UpsellRow], client: ClientInfo, order_file: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Preventivo"
    ws.append([
        "Cliente",
        client.ragione_sociale,
        "Listino",
        client.listino,
        "Ordine",
        order_file,
    ])
    ws.append([])
    ws.append([
        "Codice",
        "Descrizione",
        "Qty",
        "Prezzo (ex VAT)",
        "Ricarico %",
        "Totale (ex VAT)",
        "Disp.",
        "Disponibile dal",
    ])
    for row in rows:
        ws.append([
            row.codice,
            row.descrizione,
            row.qty,
            row.prezzo_unit,
            row.markup,
            row.totale,
            row.disp,
            row.disponibile_dal or "",
        ])
    output_path = OUTPUT_DIR / "preventivo.xlsx"
    wb.save(output_path)
    return output_path


class App(tk.Tk):
    """Main GUI for upsell computation."""

    def __init__(self) -> None:
        super().__init__()
        self.title("ORMANET UPSELLING")
        self.geometry("1100x720")
        self.configure(bg="#ffffff")
        self.logger = SessionLogger()
        self.sconti = load_json(CONFIG_DIR / "sconti_2026.json")
        self.category_map = load_json(CONFIG_DIR / "category_map.json")
        self.clients: list[ClientInfo] = []
        self.selected_client: ClientInfo | None = None
        self.stock: dict[str, StockItem] | None = None
        self.historical_orders: list[Path] = []
        self.current_order: Path | None = None
        self.upsell_rows: list[UpsellRow] = []
        self.causale = tk.StringVar(value=CAUSALI[0])
        self.aggressivita = tk.IntVar(value=0)

        self._build_ui()

    def _build_ui(self) -> None:
        header = tk.Frame(self, bg="#ff7a00", height=60)
        header.pack(fill=tk.X)
        tk.Label(
            header,
            text="ORMANET UPSELLING",
            bg="#ff7a00",
            fg="#000000",
            font=("Segoe UI", 18, "bold"),
        ).pack(side=tk.LEFT, padx=20, pady=10)

        main = tk.Frame(self, bg="#ffffff")
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        controls = tk.Frame(main, bg="#ffffff")
        controls.pack(fill=tk.X)

        self.client_combo = ttk.Combobox(controls, state="readonly", width=40)
        self.client_combo.bind("<<ComboboxSelected>>", self._on_client_selected)
        ttk.Button(controls, text="Carica clienti", command=self.load_clients_file).grid(
            row=0, column=0, padx=5, pady=5, sticky="w"
        )
        self.client_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.client_label = tk.Label(controls, text="", bg="#ffffff", fg="#000000")
        self.client_label.grid(row=0, column=2, padx=5, pady=5, sticky="w")

        ttk.Button(controls, text="Carica stock", command=self.load_stock_file).grid(
            row=1, column=0, padx=5, pady=5, sticky="w"
        )
        self.stock_label = tk.Label(controls, text="", bg="#ffffff")
        self.stock_label.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        ttk.Button(controls, text="Carica 4 storici", command=self.load_historical_files).grid(
            row=2, column=0, padx=5, pady=5, sticky="w"
        )
        self.history_label = tk.Label(controls, text="", bg="#ffffff")
        self.history_label.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        ttk.Button(controls, text="Carica ordine", command=self.load_current_file).grid(
            row=3, column=0, padx=5, pady=5, sticky="w"
        )
        self.current_label = tk.Label(controls, text="", bg="#ffffff")
        self.current_label.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        tk.Label(controls, text="Causale", bg="#ffffff").grid(row=4, column=0, sticky="w")
        ttk.Combobox(controls, textvariable=self.causale, values=CAUSALI, state="readonly").grid(
            row=4, column=1, padx=5, pady=5, sticky="w"
        )

        tk.Label(controls, text="Aggressivita Sconto", bg="#ffffff").grid(
            row=5, column=0, sticky="w"
        )
        tk.Scale(
            controls,
            from_=0,
            to=3,
            orient=tk.HORIZONTAL,
            variable=self.aggressivita,
            length=200,
            bg="#ffffff",
        ).grid(row=5, column=1, padx=5, pady=5, sticky="w")

        ttk.Button(controls, text="Calcola upsell", command=self.compute).grid(
            row=6, column=0, padx=5, pady=10, sticky="w"
        )

        self.warning_label = tk.Label(controls, text="", fg="#cc0000", bg="#ffffff")
        self.warning_label.grid(row=6, column=1, padx=5, pady=10, sticky="w")

        table_frame = tk.Frame(main, bg="#ffffff")
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        columns = (
            "codice",
            "descrizione",
            "qty",
            "prezzo",
            "markup",
            "totale",
            "disp",
            "data",
        )
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)
        for col, label in zip(
            columns,
            [
                "Codice",
                "Descrizione",
                "Qty",
                "Prezzo ex VAT",
                "Markup %",
                "Totale ex VAT",
                "Disp",
                "Disponibile dal",
            ],
        ):
            self.tree.heading(col, text=label)
            self.tree.column(col, anchor="center")
        self.tree.pack(fill=tk.BOTH, expand=True)

        actions = tk.Frame(main, bg="#ffffff")
        actions.pack(fill=tk.X, pady=10)
        ttk.Button(actions, text="Copia valori", command=self.copy_values).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(actions, text="Esporta Excel", command=self.export).pack(
            side=tk.LEFT, padx=5
        )
        ttk.Button(actions, text="Apri cartella output", command=self.open_output).pack(
            side=tk.LEFT, padx=5
        )

    def load_clients_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleziona CLIENTI.xlsx",
            initialdir=str(IMPORT_DIR),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.clients = load_clients(Path(path), self.logger)
            self.client_combo["values"] = [
                f"{client.client_id} - {client.ragione_sociale}" for client in self.clients
            ]
            self.client_label.config(text=f"{len(self.clients)} clienti")
        except Exception as exc:
            self.logger.error("Errore caricamento clienti", error=str(exc))
            messagebox.showerror("Errore", f"Errore caricamento clienti: {exc}")

    def _on_client_selected(self, _event: object) -> None:
        index = self.client_combo.current()
        if index < 0:
            return
        self.selected_client = self.clients[index]
        self.client_label.config(
            text=f"{self.selected_client.ragione_sociale} | {self.selected_client.listino}"
        )

    def load_stock_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleziona STOCK.xlsx",
            initialdir=str(IMPORT_DIR),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.stock = load_stock(Path(path), self.logger)
            self.stock_label.config(text=f"{len(self.stock)} articoli")
        except Exception as exc:
            self.logger.error("Errore caricamento stock", error=str(exc))
            messagebox.showerror("Errore", f"Errore caricamento stock: {exc}")

    def load_historical_files(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Seleziona 4 storici",
            initialdir=str(ORDERS_DIR),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not paths:
            return
        if len(paths) != 4:
            messagebox.showwarning("Attenzione", "Seleziona esattamente 4 file storici.")
            return
        self.historical_orders = [Path(path) for path in paths]
        self.history_label.config(text=f"{len(self.historical_orders)} file")

    def load_current_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleziona ordine upsell",
            initialdir=str(ORDERS_DIR),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return
        self.current_order = Path(path)
        self.current_label.config(text=self.current_order.name)

    def _validate_inputs(self) -> bool:
        if not self.selected_client:
            self.warning_label.config(text="Seleziona cliente")
            return False
        if not self.stock:
            self.warning_label.config(text="Carica stock")
            return False
        if len(self.historical_orders) != 4:
            self.warning_label.config(text="Carica 4 ordini storici")
            return False
        if not self.current_order:
            self.warning_label.config(text="Carica ordine corrente")
            return False
        self.warning_label.config(text="")
        return True

    def compute(self) -> None:
        if not self._validate_inputs():
            return
        try:
            historical_items = load_orders(self.historical_orders, self.logger)
            current_items = load_orders([self.current_order], self.logger)
            self.upsell_rows = compute_upsell(
                current_items=current_items,
                historical_items=historical_items,
                stock=self.stock or {},
                client=self.selected_client,
                sconti=self.sconti,
                category_map=self.category_map,
                aggressivita=self.aggressivita.get(),
                causale=self.causale.get(),
                logger=self.logger,
            )
            self._render_table()
        except Exception as exc:
            self.logger.error("Errore calcolo upsell", error=str(exc))
            messagebox.showerror("Errore", f"Errore calcolo upsell: {exc}")

    def _render_table(self) -> None:
        for row in self.tree.get_children():
            self.tree.delete(row)
        for row in self.upsell_rows:
            self.tree.insert(
                "",
                tk.END,
                values=(
                    row.codice,
                    row.descrizione,
                    row.qty,
                    f"{row.prezzo_unit:.2f}",
                    f"{row.markup:.2f}",
                    f"{row.totale:.2f}",
                    row.disp,
                    row.disponibile_dal or "",
                ),
            )

    def copy_values(self) -> None:
        if not self.upsell_rows:
            messagebox.showinfo("Info", "Nessuna riga upsell da copiare.")
            return
        client = self.selected_client
        order_name = self.current_order.name if self.current_order else ""
        lines = [
            f"Client: {client.ragione_sociale} (Listino: {client.listino})",
            f"Ordine: {order_name}, Causale: {self.causale.get()}",
            "Righe Upsell:",
        ]
        for row in self.upsell_rows:
            lines.append(
                f"- {row.codice} | {row.descrizione} | {row.qty} | {row.prezzo_unit:.2f} | "
                f"{row.markup:.2f}% | {row.totale:.2f} | {row.disp}"
            )
        text = "\n".join(lines)
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Copiato", "Valori copiati negli appunti.")

    def export(self) -> None:
        if not self.upsell_rows:
            messagebox.showinfo("Info", "Nessuna riga upsell da esportare.")
            return
        output_path = export_excel(
            self.upsell_rows, self.selected_client, self.current_order.name
        )
        messagebox.showinfo("Esporta", f"Export completato: {output_path}")

    def open_output(self) -> None:
        os.startfile(OUTPUT_DIR)


if __name__ == "__main__":
    app = App()
    app.mainloop()
