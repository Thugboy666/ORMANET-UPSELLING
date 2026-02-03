"""XLSX loading helpers for Ormanet Upselling."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from app.engine import ClientInfo, OrderItem, SessionLogger, StockItem

DEFAULT_FIELD_MAPPING: dict[str, dict[str, list[str]]] = {
    "ORDINI": {
        "marca": ["Marca"],
        "categoria": ["Categoria"],
        "codice": ["Cod.", "Cod", "Codice"],
        "descrizione": ["Descrizione"],
        "qty": ["Qty", "Qta", "Q.tà", "Quantità", "Quantita"],
        "mag_centr": [
            "Mag. Centr.",
            "Mag. Centr",
            "Mag.Centr.",
            "Disp.",
            "Disponibilità",
            "Disponibilita",
        ],
        "mag_in_arr": [
            "Mag. in Arr.",
            "Mag. in Arr",
            "Disp. in Arrivo",
            "Disponibilità in Arrivo",
            "Disponibilita in Arrivo",
        ],
        "evaso": ["Evaso"],
        "data_arrivo": ["Data arrivo", "Data evasione/arrivo", "Disponibile dal"],
        "prezzo_unit_exvat": [
            "Prezzo",
            "Pr.sc.",
            "Prezzo unit",
            "Prezzo (unit, ex VAT)",
            "Prezzo unit (ex VAT)",
        ],
        "totale_exvat": ["Totale", "Imponibile", "Totale (ex VAT)"],
        "s1": ["S1"],
        "s2": ["S2"],
        "listino_ri10": ["L-RI10", "Listino RI+10%", "Listino RI+10"],
        "listino_ri": ["L-RI", "Listino RI"],
        "listino_di": ["L-DI", "Listino DI"],
        "lm": ["LM", "Listino madre", "Listino Madre"],
    },
    "STOCK": {
        "categoria": ["Categoria"],
        "marca": ["Marca"],
        "codice": ["Codice", "Cod.", "Cod"],
        "descrizione": ["Descrizione"],
        "disp": ["Disp.", "Disponibile", "Disponibilita"],
        "disp_in_arrivo": [
            "Disp. in Arrivo",
            "Disponibilità in Arrivo",
            "Disponibilita in Arrivo",
            "Mag. in Arr.",
        ],
        "giacenza": ["Giacenza"],
        "data_evasione_arrivo": ["Data evasione/arrivo", "Data arrivo"],
        "listino_ri10": ["Listino RI+10%", "Listino RI+10"],
        "listino_ri": ["Listino RI"],
        "listino_di": ["Listino DI"],
        "lm": ["LM", "Listino madre", "Listino Madre"],
        "prezzo_alt": ["PREZZO_ALT", "Prezzo Alt", "Prezzo ALT"],
    },
    "CLIENTI": {
        "id": ["ID", "Codice", "Codice numerico", "Cod. Cliente"],
        "ragione_sociale": ["Ragione sociale", "Cliente", "Nominativo"],
        "zona": ["Zona", "Regione"],
        "listino": ["Listino"],
        "categoria_listino": ["Categoria"],
        "note_consegna": ["Note consegna"],
        "note_amministrative": ["Note amministrative"],
        "note": ["Note"],
        "email": ["Email", "E-mail"],
        "telefono": ["Telefono"],
        "cellulare": ["Cellulare"],
        "pagamento": ["Pagamento"],
        "fatturato": ["Fatturato"],
    },
}

REQUIRED_FIELDS = {
    "ORDINI": ["codice", "qty", "prezzo_unit_exvat"],
    "STOCK": ["codice", "disp"],
    "CLIENTI": ["id", "ragione_sociale", "listino"],
}

STOCK_LISTINO_FIELDS = ["listino_ri", "listino_ri10", "listino_di"]


class MappingError(ValueError):
    def __init__(self, message: str, details: dict[str, Any]) -> None:
        super().__init__(message)
        self.details = details


class DataError(ValueError):
    def __init__(self, message: str, details: dict[str, Any]) -> None:
        super().__init__(message)
        self.details = details


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u00a0", " ")
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\.+$", "", text)
    text = text.replace(" %", "%").replace("% ", "%")
    return text


def normalize_mapping(mapping: dict[str, dict[str, list[str]]]) -> dict[str, dict[str, list[str]]]:
    return deepcopy(mapping)


def read_headers(path: Path) -> list[str]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    return ["" if value is None else str(value).strip() for value in raw_headers]


def get_cell(row: tuple[Any, ...], idx: int | None, default: Any = "") -> Any:
    if idx is None:
        return default
    if idx >= len(row):
        return default
    return row[idx]


def build_header_map(headers: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, name in enumerate(headers):
        normalized = normalize_header(name)
        if normalized and normalized not in header_map:
            header_map[normalized] = idx
    return header_map


def match_mapping(
    headers: list[str],
    mapping: dict[str, list[str]],
) -> tuple[dict[str, str | None], dict[str, int | None]]:
    header_map = build_header_map(headers)
    matches: dict[str, str | None] = {}
    indices: dict[str, int | None] = {}
    for field, aliases in mapping.items():
        matched_idx = None
        matched_header: str | None = None
        for alias in aliases:
            alias_norm = normalize_header(alias)
            if alias_norm in header_map:
                matched_idx = header_map[alias_norm]
                matched_header = headers[matched_idx]
                break
        matches[field] = matched_header
        indices[field] = matched_idx
    return matches, indices


def parse_float(
    value: Any,
    field_name: str,
    row_index: int | None = None,
    filename: str | None = None,
) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return 0.0
    cleaned = text.replace("€", "").replace(" ", "").replace("\u00a0", "")
    cleaned = cleaned.replace("%", "")
    if "." in cleaned and "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError as exc:
        details = {
            "field": field_name,
            "raw_value": text,
            "row": row_index,
            "file": filename,
        }
        raise DataError(
            f"Valore non numerico per '{field_name}': '{text}'",
            details,
        ) from exc


def parse_optional_price(
    value: Any,
    field_name: str,
    row_index: int | None,
    filename: str | None,
    logger: SessionLogger,
) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = parse_float(value, field_name, row_index, filename)
    except DataError as exc:
        logger.info(
            "Valore non valido per %s in %s riga %s: %s",
            field_name,
            filename or "-",
            row_index or "-",
            exc.details.get("raw_value"),
        )
        return None
    if parsed < 0:
        logger.info(
            "Valore negativo per %s in %s riga %s: %.2f",
            field_name,
            filename or "-",
            row_index or "-",
            parsed,
        )
        return None
    return parsed


def _raise_mapping_error(
    logger: SessionLogger,
    mapping_type: str,
    missing_fields: list[str],
    headers: list[str],
    path: Path,
) -> None:
    details = {
        "mapping_type": mapping_type,
        "missing_fields": missing_fields,
        "found_headers": headers,
        "file": str(path),
    }
    logger.error("Mapping error", error_type="mapping_error", details=details)
    raise MappingError(
        f"Campi richiesti mancanti ({mapping_type}): {', '.join(missing_fields)}",
        details,
    )


def _require_fields(
    logger: SessionLogger,
    mapping_type: str,
    indices: dict[str, int | None],
    headers: list[str],
    path: Path,
) -> None:
    missing = [field for field in REQUIRED_FIELDS[mapping_type] if indices.get(field) is None]
    if mapping_type == "STOCK":
        has_listino = any(indices.get(field) is not None for field in STOCK_LISTINO_FIELDS)
        if not has_listino:
            missing.append("listino_ri|listino_ri10|listino_di")
    if missing:
        _raise_mapping_error(logger, mapping_type, missing, headers, path)


def load_clients(
    path: Path,
    logger: SessionLogger,
    mapping: dict[str, list[str]],
) -> list[ClientInfo]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    headers = ["" if value is None else str(value).strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    matches, indices = match_mapping(headers, mapping)
    logger.info("CLIENTI headers (%s): %s", path.name, headers)
    logger.info("CLIENTI mapping matches (%s): %s", path.name, matches)
    _require_fields(logger, "CLIENTI", indices, headers, path)

    clients: list[ClientInfo] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        client_id = str(get_cell(row, indices.get("id")) or "").strip()
        ragione_sociale = str(get_cell(row, indices.get("ragione_sociale")) or "").strip()
        listino = str(get_cell(row, indices.get("listino")) or "").strip()
        categoria = str(get_cell(row, indices.get("categoria_listino")) or "").strip()
        if not client_id or not ragione_sociale:
            continue
        clients.append(ClientInfo(client_id, ragione_sociale, listino, categoria))
    logger.info("Loaded %s clients", len(clients))
    return clients


def load_stock(
    path: Path,
    logger: SessionLogger,
    mapping: dict[str, list[str]],
) -> dict[str, StockItem]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    headers = ["" if value is None else str(value).strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    matches, indices = match_mapping(headers, mapping)
    logger.info("STOCK headers (%s): %s", path.name, headers)
    logger.info("STOCK mapping matches (%s): %s", path.name, matches)
    _require_fields(logger, "STOCK", indices, headers, path)

    stock: dict[str, StockItem] = {}
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        codice = str(get_cell(row, indices.get("codice")) or "").strip()
        if not codice:
            continue
        stock[codice] = StockItem(
            categoria=str(get_cell(row, indices.get("categoria")) or "").strip(),
            marca=str(get_cell(row, indices.get("marca")) or "").strip(),
            codice=codice,
            descrizione=str(get_cell(row, indices.get("descrizione")) or "").strip(),
            disp=parse_float(get_cell(row, indices.get("disp"), 0), "disp", row_index, path.name),
            disp_in_arrivo=parse_float(
                get_cell(row, indices.get("disp_in_arrivo"), 0),
                "disp_in_arrivo",
                row_index,
                path.name,
            ),
            giacenza=parse_float(
                get_cell(row, indices.get("giacenza"), 0),
                "giacenza",
                row_index,
                path.name,
            ),
            data_arrivo=str(get_cell(row, indices.get("data_evasione_arrivo")) or "").strip(),
            listino_ri10=parse_float(
                get_cell(row, indices.get("listino_ri10"), 0),
                "listino_ri10",
                row_index,
                path.name,
            ),
            listino_ri=parse_float(
                get_cell(row, indices.get("listino_ri"), 0),
                "listino_ri",
                row_index,
                path.name,
            ),
            listino_di=parse_float(
                get_cell(row, indices.get("listino_di"), 0),
                "listino_di",
                row_index,
                path.name,
            ),
            lm=parse_float(get_cell(row, indices.get("lm"), 0), "lm", row_index, path.name),
            prezzo_alt=parse_optional_price(
                get_cell(row, indices.get("prezzo_alt")),
                "prezzo_alt",
                row_index,
                path.name,
                logger,
            ),
            source_file=path.name,
            source_row=row_index,
        )
    logger.info("Loaded %s stock items", len(stock))
    return stock


def load_orders(
    paths: list[Path],
    logger: SessionLogger,
    mapping: dict[str, list[str]],
) -> list[OrderItem]:
    items: list[OrderItem] = []
    for path in paths:
        wb = load_workbook(filename=path, data_only=True)
        ws = wb.active
        headers = ["" if value is None else str(value).strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        matches, indices = match_mapping(headers, mapping)
        logger.info("ORDINI headers (%s): %s", path.name, headers)
        logger.info("ORDINI mapping matches (%s): %s", path.name, matches)
        _require_fields(logger, "ORDINI", indices, headers, path)
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            codice = str(get_cell(row, indices.get("codice")) or "").strip()
            if not codice:
                continue
            items.append(
                OrderItem(
                    marca=str(get_cell(row, indices.get("marca")) or "").strip(),
                    categoria=str(get_cell(row, indices.get("categoria")) or "").strip(),
                    codice=codice,
                    descrizione=str(get_cell(row, indices.get("descrizione")) or "").strip(),
                    qty=parse_float(get_cell(row, indices.get("qty"), 0), "qty", row_index, path.name),
                    prezzo_unit=parse_float(
                        get_cell(row, indices.get("prezzo_unit_exvat"), 0),
                        "prezzo_unit_exvat",
                        row_index,
                        path.name,
                    ),
                    lm=parse_float(get_cell(row, indices.get("lm"), 0), "lm", row_index, path.name),
                    source_file=path.name,
                    source_row=row_index,
                )
            )
    logger.info("Loaded %s order items", len(items))
    return items
